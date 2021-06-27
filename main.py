import copy
import openpyxl
import datetime
import logging
from my_func import *
from alpha_index import *
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

try:
    # path_config = input('Укажите путь до конфигуратора\n')
    with open('Source_list_config', 'r', encoding='UTF-8') as f:
        path_config = f.readline().strip()

    print(datetime.datetime.now())

    file_config = 'UnimodCreate.xlsm'

    # Ищем файл конфигуратора в указанном каталоге
    for file in os.listdir(path_config):
        if file.endswith('.xlsm') or file.endswith('.xls'):
            file_config = file
            break

    all_CPU = []
    sl_CPU_spec = {}
    pref_IP = []
    sl_object_rus = {}
    sl_object_all = {}
    num_pz = 0
    lst_all_wrn = []
    lst_all_ts = []
    lst_all_ppu = []
    lst_all_mod = []
    lst_all_alg = []
    sl_all_pz = {}
    sl_for_diag = {}
    sl_cpu_drv_signal = {}

    # Считываем файл-шаблон для AI  AE SET
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_AIAESET'), 'r', encoding='UTF-8') as f:
        tmp_object_AIAESET = f.read()
    # Считываем файл-шаблон для DI
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_DI'), 'r', encoding='UTF-8') as f:
        tmp_object_DI = f.read()
    # Считываем файл-шаблон для IM
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_IM'), 'r', encoding='UTF-8') as f:
        tmp_object_IM = f.read()
    # Считываем файл-шаблон для BTN CNT
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_BTN_CNT_sig'), 'r', encoding='UTF-8') as f:
        tmp_object_BTN_CNT_sig = f.read()  # его же используем для диагностики CPU потому что подходит
    # Считываем файл-шаблон для PZ
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_PZ'), 'r', encoding='UTF-8') as f:
        tmp_object_PZ = f.read()
    # Считываем файл-шаблон для группы
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_group'), 'r', encoding='UTF-8') as f:
        tmp_group = f.read()
    # Считываем файл-шаблон для app
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_app'), 'r', encoding='UTF-8') as f:
        tmp_app = f.read()
    # Считываем файл-шаблон для контроллера
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_TREI'), 'r', encoding='UTF-8') as f:
        tmp_trei = f.read()
    # Считываем файл-шаблон для global
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_global'), 'r', encoding='UTF-8') as f:
        tmp_global = f.read()
    # Считываем файл-шаблон для топливного регулятора
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_TR_ps90'), 'r', encoding='UTF-8') as f:
        tmp_tr_ps90 = f.read()
    # Считываем файл-шаблон для АПР
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_APR'), 'r', encoding='UTF-8') as f:
        tmp_apr = f.read()
    # Считываем файл-шаблон для драйверных параметров
    with open(os.path.join(os.path.dirname(__file__), 'Template', 'Temp_drv_par'), 'r', encoding='UTF-8') as f:
        tmp_drv_par = f.read()

    book = openpyxl.open(os.path.join(path_config, file_config))  # , read_only=True
    # читаем список всех контроллеров
    sheet = book['Настройки']  # worksheets[1]
    cells = sheet['B2': 'B22']
    for p in cells:
        if p[0].value is not None:
            all_CPU.append(p[0].value)
    # Читаем префиксы IP адреса ПЛК(нужно продумать про новые конфигураторы)
    cells = sheet['B45': 'B46']
    for p in cells:
        pref_IP.append(p[0].value + '.')

    # Читаем состав объектов
    cells = sheet['B24': 'R38']
    for p in cells:
        if p[0].value is not None:
            sl_object_rus[p[0].value] = p[1].value
            tmp0 = []
            tmp1 = []
            for i in range(12, 17):
                if p[i].value == 'ON':
                    tmp0.append(p[i - 10].value)
                    tmp1.append(p[i - 5].value)
            sl_object_all[p[0].value] = [tmp0, tmp1]
    # Мониторинг ТР и АПР в контроллере
    cells = sheet['B1':'L21']
    for p in cells:
        if p[0].value is None:
            break
        else:
            sl_CPU_spec[p[0].value] = []
            if p[is_f_ind(cells[0], 'FLR')].value == 'ON' and p[is_f_ind(cells[0], 'Тип ТР')].value == 'ПС90':
                sl_CPU_spec[p[0].value].append('ТР')
            if p[is_f_ind(cells[0], 'APR')].value == 'ON':
                sl_CPU_spec[p[0].value].append('АПР')

    # Определение заведённых драйверов
    cells = sheet['A1': 'A' + str(sheet.max_row)]
    drv_eng, drv_rus = [], []
    for p in cells:
        if p[0].value == 'Наименование драйвера (Eng)':
            jj = 1
            while sheet[p[0].row][jj].value and sheet[p[0].row + 1][jj].value:
                drv_eng.append(sheet[p[0].row][jj].value)
                drv_rus.append(sheet[p[0].row + 1][jj].value)
                jj += 1
    sl_all_drv = dict(zip(drv_eng, drv_rus))
    ff = open('file_plc.txt', 'w', encoding='UTF-8')
    ff.close()
    # Далее для всех контроллеров, что нашли, делаем
    for i in all_CPU:
        # Диагностика модулей (DIAG)
        sheet = book['Модули']
        cells = sheet['A1': 'G' + str(sheet.max_row)]
        sl_modules_cpu = {}
        for p in cells:
            if p[is_f_ind(cells[0], 'CPU')].value == i:
                aa = copy.copy(sl_modules[p[is_f_ind(cells[0], 'Шифр модуля')].value])
                sl_modules_cpu[p[is_f_ind(cells[0], 'Имя модуля')].value] = [p[is_f_ind(cells[0], 'Шифр модуля')].value,
                                                                             aa]

        for jj in ['Измеряемые', 'Входные', 'Выходные', 'ИМ(АО)']:
            sheet_run = book[jj]
            cells_run = sheet_run['A1': 'O' + str(sheet_run.max_row)]
            for p in cells_run:
                if p[is_f_ind(cells_run[0], 'CPU')].value == i and \
                        p[is_f_ind(cells_run[0], 'Нестандартный канал')].value == 'Нет':
                    tmp_ind = int(p[is_f_ind(cells_run[0], 'Номер канала')].value) - 1
                    sl_modules_cpu[p[is_f_ind(cells_run[0], 'Номер модуля')].value][1][tmp_ind] = \
                        is_cor_chr(p[is_f_ind(cells_run[0], 'Наименование параметра')].value)

        if len(sl_modules_cpu) != 0:
            tmp_line_ = is_create_objects_diag(sl_modules_cpu, tmp_object_BTN_CNT_sig)
            tmp_line_ = (Template(tmp_group).substitute(name_group='HW', objects=tmp_line_))

            with open('file_out_group.txt', 'w', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='Diag', objects=tmp_line_.rstrip()))

            # для каждого контроллера сначала создаём пустой словарь в словаре
            sl_for_diag[i] = {}
            # далее проходим по локальному словарю и для каждого модуля грузим {алг.имя модуля: тип модуля}
            for jj in sl_modules_cpu:
                # в случае CPU - {CPU: алг.имя модуля}
                if sl_modules_cpu[jj][0] in ('M903E', 'M991E'):
                    sl_for_diag[i].update({'CPU': jj})
                else:
                    sl_for_diag[i].update({jj: sl_modules_cpu[jj][0]})

        # Измеряемые
        sheet = book['Измеряемые']  # .worksheets[3]
        cells = sheet['A1': 'AG' + str(sheet.max_row)]
        sl_CPU_one = is_load_ai_ae_set(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                       is_f_ind(cells[0], 'Наименование параметра'),
                                       is_f_ind(cells[0], 'Единицы измерения'),
                                       is_f_ind(cells[0], 'Короткое наименование'),
                                       is_f_ind(cells[0], 'Количество знаков'),
                                       is_f_ind(cells[0], 'CPU'))

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_ai_ae_set(sl_CPU_one, tmp_object_AIAESET, 'Types.AI.AI_PLC_View')

            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='AI', objects=tmp_line_))

        # Расчетные
        sheet = book['Расчетные']  # .worksheets[4]
        cells = sheet['A1': 'AE' + str(sheet.max_row)]
        sl_CPU_one = is_load_ai_ae_set(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                       is_f_ind(cells[0], 'Наименование параметра'),
                                       is_f_ind(cells[0], 'Единицы измерения'),
                                       is_f_ind(cells[0], 'Короткое наименование'),
                                       is_f_ind(cells[0], 'Количество знаков'),
                                       is_f_ind(cells[0], 'CPU'))

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_ai_ae_set(sl_CPU_one, tmp_object_AIAESET, 'Types.AE.AE_PLC_View')

            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='AE', objects=tmp_line_))

        # Дискретные
        sheet = book['Входные']  # .worksheets[6]
        cells = sheet['A1': 'AC' + str(sheet.max_row)]
        sl_CPU_one, sl_wrn = is_load_di(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                        is_f_ind(cells[0], 'ИМ'),
                                        is_f_ind(cells[0], 'Наименование параметра'),
                                        is_f_ind(cells[0], 'Цвет при наличии'),
                                        is_f_ind(cells[0], 'Цвет при отсутствии'),
                                        is_f_ind(cells[0], 'Предупреждение'),
                                        is_f_ind(cells[0], 'Текст предупреждения'),
                                        is_f_ind(cells[0], 'CPU'))

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_di(sl_CPU_one, tmp_object_DI, 'Types.DI.DI_PLC_View')

            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='DI', objects=tmp_line_))

        # ИМ
        sheet = book['ИМ']  # .worksheets[9]
        cells = sheet['A1': 'T' + str(sheet.max_row)]
        sl_CPU_one = is_load_im(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                is_f_ind(cells[0], 'Наименование параметра'),
                                is_f_ind(cells[0], 'Тип ИМ'),
                                is_f_ind(cells[0], 'Род'),
                                is_f_ind(cells[0], 'Считать наработку'),
                                is_f_ind(cells[0], 'Считать перестановки'),
                                is_f_ind(cells[0], 'CPU'))
        sl_cnt = {}
        for key, value in sl_CPU_one.items():
            if value[4] == 'Да':
                sl_cnt[key+'_WorkTime'] = [value[0]]
            if value[5] == 'Да':
                sl_cnt[key+'_Swap'] = [value[0]]

        # ИМ АО- объединяем словари с ИМами
        sheet = book['ИМ(АО)']  # .worksheets[8]
        cells = sheet['A1': 'AA' + str(sheet.max_row)]
        sl_CPU_one = {**sl_CPU_one, **is_load_im_ao(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                                    is_f_ind(cells[0], 'Наименование параметра'),
                                                    is_f_ind(cells[0], 'Род'),
                                                    is_f_ind(cells[0], 'ИМ'),
                                                    is_f_ind(cells[0], 'CPU'))}

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_im(sl_CPU_one, tmp_object_IM)

            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='IM', objects=tmp_line_))

        # Добавляем АПР, если для данного контроллера указан АПР в настройках
        if 'АПР' in sl_CPU_spec[i]:
            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(tmp_apr.rstrip())

        # Кнопки(в составе System)
        sheet = book['Кнопки']  # .worksheets[10]
        cells = sheet['A1': 'C' + str(sheet.max_row)]
        sl_CPU_one = is_load_btn(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                 is_f_ind(cells[0], 'Наименование параметра'),
                                 is_f_ind(cells[0], 'CPU'))

        tmp_subgroup = ''
        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_btn_cnt(sl_CPU_one, tmp_object_BTN_CNT_sig, 'Types.BTN.BTN_PLC_View')
            tmp_subgroup += Template(tmp_group).substitute(name_group='BTN', objects=tmp_line_)

        # Уставки(в составе System)
        sheet = book['Уставки']  # .worksheets[5]
        cells = sheet['A1': 'AG' + str(sheet.max_row)]
        sl_CPU_one = is_load_ai_ae_set(i, cells, is_f_ind(cells[0], 'Алгоритмическое имя'),
                                       is_f_ind(cells[0], 'Наименование параметра'),
                                       is_f_ind(cells[0], 'Единицы измерения'),
                                       is_f_ind(cells[0], 'Короткое наименование'),
                                       is_f_ind(cells[0], 'Количество знаков'),
                                       is_f_ind(cells[0], 'CPU'))

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_ai_ae_set(sl_CPU_one, tmp_object_AIAESET, 'Types.SET.SET_PLC_View')
            tmp_subgroup += Template(tmp_group).substitute(name_group='SET', objects=tmp_line_)

        # Наработки(CNT- прочитали при ИМ) в составе System)
        if len(sl_cnt) != 0:
            tmp_line_ = is_create_objects_btn_cnt(sl_cnt, tmp_object_BTN_CNT_sig, 'Types.CNT.CNT_PLC_View')
            tmp_subgroup += Template(tmp_group).substitute(name_group='CNT', objects=tmp_line_)

        # Защиты(PZ) в составе System
        sheet = book['Сигналы']  # .worksheets[11]
        cells = sheet['A1': 'N' + str(sheet.max_row)]
        sl_all_pz[i] = [num_pz]
        sl_CPU_one, num_pz = is_load_pz(i, cells, num_pz,
                                        is_f_ind(cells[0], 'Наименование параметра'),
                                        is_f_ind(cells[0], 'Тип защиты'),
                                        is_f_ind(cells[0], 'Единица измерения'),
                                        is_f_ind(cells[0], 'CPU'))
        if num_pz not in sl_all_pz[i]:
            sl_all_pz[i].append(num_pz)

        if len(sl_CPU_one) != 0:
            tmp_line_ = is_create_objects_pz(sl_CPU_one, tmp_object_PZ, 'Types.PZ.PZ_PLC_View')
            tmp_subgroup += Template(tmp_group).substitute(name_group='PZ', objects=tmp_line_)

        # ПС(WRN) в составе System, каждая ПС как отдельный объект, дополняем словарь, созданный при анализе DI
        tmp_wrn, sl_ts, sl_ppu, sl_alr, sl_modes, sl_alg = is_load_sig(i, cells,
                                                                       is_f_ind(cells[0], 'Алгоритмическое имя'),
                                                                       is_f_ind(cells[0], 'Наименование параметра'),
                                                                       is_f_ind(cells[0], 'Тип защиты'),
                                                                       is_f_ind(cells[0], 'CPU'))
        sl_wrn = {**sl_wrn, **tmp_wrn}

        if len(sl_wrn) != 0:
            tmp_line_ = is_create_objects_sig(sl_wrn, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='WRN', objects=tmp_line_)
            lst_all_wrn += list(sl_wrn.keys())

        # ТС в составе System, сам словарь загружен ранее вместе с ПС
        if len(sl_ts) != 0:
            tmp_line_ = is_create_objects_sig(sl_ts, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='TS', objects=tmp_line_)
            lst_all_ts += list(sl_ts.keys())

        # ППУ в составе System, сам словарь загружен ранее вместе с ПС
        if len(sl_ppu) != 0:
            tmp_line_ = is_create_objects_sig(sl_ppu, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='PPU', objects=tmp_line_)
            lst_all_ppu += list(sl_ppu.keys())

        # ALR(Защиты и АС) в составе System, сам словарь загружен ранее вместе с ПС
        if len(sl_alr) != 0:
            tmp_line_ = is_create_objects_sig(sl_alr, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='ALR', objects=tmp_line_)

        # MODES(Режимы) в составе System, сам словарь загружен ранее вместе с ПС
        if len(sl_modes) != 0:
            tmp_line_ = is_create_objects_sig(sl_modes, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='MODES', objects=tmp_line_)
            lst_all_mod += list(sl_modes.keys())

        # ALG в составе System, сам словарь загружен ранее вместе с ПС(позже поддержать новый конфигуратор)
        if len(sl_alg) != 0:
            tmp_line_ = is_create_objects_sig(sl_alg, tmp_object_BTN_CNT_sig)
            tmp_subgroup += Template(tmp_group).substitute(name_group='ALG', objects=tmp_line_)
            lst_all_alg += list(sl_alg.keys())

        # Добавляем топливный регулятор, если для данного контроллера указан ТР ПС90 в настройках
        if 'ТР' in sl_CPU_spec[i]:
            tmp_subgroup += tmp_tr_ps90.rstrip()

        # Драйвера в составе System
        sheet = book['Драйвера']
        cells = sheet['A1': 'N' + str(sheet.max_row)]
        '''
        for jj in range(len(cells[0])):
            print(jj, multiple_replace(cells[0][jj].value))  #
        '''
        sl_CPU_one = is_load_drv(controller=i, cell=cells, alg_name=is_f_ind(cells[0], 'Алгоритмическое имя'),
                                 name_par=is_f_ind(cells[0], 'Наименование параметра'),
                                 eunit=is_f_ind(cells[0], 'Единица измерения'),
                                 type_sig=is_f_ind(cells[0], 'Тип'),
                                 type_msg=is_f_ind(cells[0], 'Тип сообщения'),
                                 c_on=is_f_ind(cells[0], 'Цвет при наличии'),
                                 c_off=is_f_ind(cells[0], 'Цвет при отсутствии'),
                                 f_dig=is_f_ind(cells[0], 'Число знаков'),
                                 cpu=is_f_ind(cells[0], 'CPU'))
        if sl_CPU_one:
            sl_drv_cpu = {}  # {(алг.имя драйвера, русское наименование драйвера): [сигналы драйвера с обвесами]}
            for key in sl_CPU_one:
                if key[0] in sl_all_drv:
                    v = (key[0], sl_all_drv[key[0]])
                    if v not in sl_drv_cpu:
                        sl_drv_cpu[v] = [(key[1], *sl_CPU_one[key])]
                    else:
                        sl_drv_cpu[v].append((key[1], *sl_CPU_one[key]))
            # print(sl_drv)
            # print(sl_drv_cpu[('DRV_EF_Rtg', 'Расходомер ТГ')])
            # в общем словаре сигналов драйверов создаём пустой словарь с текущим контроллером
            sl_cpu_drv_signal[i] = {}
            for key in sl_drv_cpu:
                # для каждого драйвера создаём пустой кортеж
                tmp_sl_plus = ()
                for value in sl_drv_cpu[key]:  # пробегам по сигналам драйвера
                    # в созданный ранее кортеж добавляем алг. имя переменной
                    tmp_sl_plus += (value[0],)
                # после создания кортежа со всеми переменными драйвера
                # обновляем общий словарь с ключом текущего контроллера(i) словарём
                # {алг. имя драйвера: (кортеж его переменных)}
                sl_cpu_drv_signal[i].update({key[0]: tmp_sl_plus})

            tmp_sub_drv = ''
            for drv in sl_drv_cpu:
                tmp_line_ = is_create_objects_drv(sl_drv_cpu=sl_drv_cpu, tuple_name_drv=drv,
                                                  template_text=tmp_drv_par)
                tmp_sub_drv += Template(tmp_group).substitute(name_group=drv[0], objects=tmp_line_)

            tmp_subgroup += Template(tmp_group).substitute(name_group='DRV', objects=tmp_sub_drv)

        # Формируем подгруппу
        if tmp_subgroup != '':
            with open('file_out_group.txt', 'a', encoding='UTF-8') as f:
                f.write(Template(tmp_group).substitute(name_group='System', objects=tmp_subgroup.rstrip()))

        # Формирование выходного файла app
        with open('file_out_group.txt', 'r', encoding='UTF-8') as f:
            tmp_line_ = f.read().rstrip()
            tmp_line_ += '\n<trei:unet-address-map name="UnetAddressMap" />\n'
        with open('file_app_out.txt', 'w', encoding='UTF-8') as f:
            f.write(Template(tmp_app).substitute(name_app='Tree', ct_object=tmp_line_))

        with open('file_app_out.txt', 'r', encoding='UTF-8') as f:
            tmp_line_ = f.read().rstrip()

        # print(os.path.exists(os.path.join(os.path.dirname(__file__), 'File_out')))
        # Если нет папки File_out, то создадим её
        if not os.path.exists(os.path.join(os.path.dirname(__file__), 'File_out')):
            os.mkdir(os.path.join(os.path.dirname(__file__), 'File_out'))
        if not os.path.exists(os.path.join(os.path.dirname(__file__), 'File_out', 'PLC_Aspect_importDomain')):
            os.mkdir(os.path.join(os.path.dirname(__file__), 'File_out', 'PLC_Aspect_importDomain'))
        # Для каждого объекта создаём контроллер
        num_obj_plc = 1
        for obj in sl_object_all:
            if i in sl_object_all[obj][0]:
                index_tmp = sl_object_all[obj][0].index(i)
                with open('file_plc.txt', 'a', encoding='UTF-8') as f:
                    f.write(Template(tmp_trei).substitute(plc_name='PLC_' + i + str(num_obj_plc), plc_name_type='m903e',
                                                          ip_eth1=pref_IP[0] + sl_object_all[obj][1][index_tmp],
                                                          ip_eth2=pref_IP[1] + sl_object_all[obj][1][index_tmp],
                                                          dp_app=tmp_line_))
                # Для каждого контроллера создадим отдельный файл для импорта только его одного
                tmp_plc = Template(tmp_trei).substitute(plc_name='PLC_' + i + str(num_obj_plc), plc_name_type='m903e',
                                                        ip_eth1=pref_IP[0] + sl_object_all[obj][1][index_tmp],
                                                        ip_eth2=pref_IP[1] + sl_object_all[obj][1][index_tmp],
                                                        dp_app=tmp_line_)
                # Если в выходной папке ПЛК-аспектов уже есть формируемый файл
                if os.path.exists(os.path.join(os.path.dirname(__file__), 'File_out', 'PLC_Aspect_importDomain',
                                               f'file_out_plc_{i}_{num_obj_plc}.omx-export')):
                    # то формируем новую версию файла
                    new_tmp_plc = Template(tmp_global).substitute(dp_node=tmp_plc)
                    # считываем имеющейся файл
                    with open(os.path.join(os.path.dirname(__file__), 'File_out', 'PLC_Aspect_importDomain',
                                           f'file_out_plc_{i}_{num_obj_plc}.omx-export'),
                              'r', encoding='UTF-8') as f_check:
                        tmp_plc_check = f_check.read()
                    # print('YES' if new_tmp_plc == tmp_plc_check else 'NO')

                    # Если новый и старый файл отличаются
                    if new_tmp_plc != tmp_plc_check:
                        # Если нет папки Old, то создаём её
                        if not os.path.exists(os.path.join(os.path.dirname(__file__), 'File_out',
                                                           'PLC_Aspect_importDomain', 'Old')):
                            os.mkdir(os.path.join(os.path.dirname(__file__), 'File_out',
                                                  'PLC_Aspect_importDomain', 'Old'))
                        # Переносим старую файл в папку Old
                        os.replace(os.path.join(os.path.dirname(__file__), 'File_out',
                                                'PLC_Aspect_importDomain',
                                                f'file_out_plc_{i}_{num_obj_plc}.omx-export'),
                                   os.path.join(os.path.dirname(__file__), 'File_out',
                                                'PLC_Aspect_importDomain', 'Old',
                                                f'file_out_plc_{i}_{num_obj_plc}.omx-export'))
                        # Записываем новый файл
                        with open(os.path.join(os.path.dirname(__file__), 'File_out', 'PLC_Aspect_importDomain',
                                               f'file_out_plc_{i}_{num_obj_plc}.omx-export'),
                                  'w', encoding='UTF-8') as f:
                            f.write(new_tmp_plc)
                        # пишем, что надо заменить
                        print(f'Требуется заменить карту ПЛК-аспект {i}_{num_obj_plc}')
                # Если в выходной папке ПЛК-аспектов нет формируемого файла, то создаём его и пишем, что заменить
                else:
                    with open(os.path.join(os.path.dirname(__file__), 'File_out', 'PLC_Aspect_importDomain',
                                           f'file_out_plc_{i}_{num_obj_plc}.omx-export'),
                              'w', encoding='UTF-8') as f:
                        f.write(Template(tmp_global).substitute(dp_node=tmp_plc))
                    print(f'Требуется заменить карту ПЛК-аспект {i}_{num_obj_plc}')
                # прибавляем номер объекта для формирования следующего файла
                num_obj_plc += 1
            else:
                num_obj_plc += 1
                continue

        # чистим файл групп для корректной обработки отсуствия
        file_clear = open('file_out_group.txt', 'w', encoding='UTF-8')
        file_clear.close()

    # Формирование выходного файла для каждого контроллера
    with open('file_plc.txt', 'r', encoding='UTF-8') as f:
        tmp_line_ = f.read().rstrip()

    with open('file_out_plc_aspect.omx-export', 'w', encoding='UTF-8') as f:
        f.write(Template(tmp_global).substitute(dp_node=tmp_line_))

    book.close()

    os.remove('file_plc.txt')
    os.remove('file_out_group.txt')
    # os.remove('file_out_objects.txt')
    os.remove('file_app_out.txt')
    print(datetime.datetime.now())
    # for k in sl_for_diag:
    #    print(k, sl_for_diag[k])
    # print()
    # for k in sl_cpu_drv_signal:
    #    print(k, sl_cpu_drv_signal[k])
    # print()
    create_index(lst_alg=lst_all_alg, lst_mod=lst_all_mod, lst_ppu=lst_all_ppu, lst_ts=lst_all_ts, lst_wrn=lst_all_wrn,
                 sl_pz_anum=sl_all_pz, sl_cpu_spec=sl_CPU_spec, sl_diag=sl_for_diag,
                 sl_cpu_drv_signal=sl_cpu_drv_signal)
    print(datetime.datetime.now())

except (Exception, KeyError):
    # в случае возникновения какой-либо ошибки, чистим возможные промежуточные файлы
    for file_error_clear in ('file_plc.txt', 'file_out_group.txt', 'file_app_out.txt', 'file_app_out.txt'):
        if os.path.exists(file_error_clear):
            os.remove(file_error_clear)

    logging.basicConfig(filename='error.log', filemode='a', datefmt='%d.%m.%y %H:%M:%S',
                        format='%(levelname)s - %(message)s - %(asctime)s')
    logging.exception("Ошибка выполнения")
    print('Произошла ошибка выполнения')
